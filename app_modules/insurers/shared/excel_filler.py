from copy import copy
from io import BytesIO

import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Color, PatternFill

from app_modules.insurers.shared.sheet_config import SHEET_MAPPINGS, transform_for_sheet

HEADLINE_COLORS = ["FF0BD7B5", "0BD7B5"]


def _is_headline_cell(cell) -> bool:
    fill = cell.fill
    return bool(
        fill
        and isinstance(fill, PatternFill)
        and fill.fgColor
        and fill.fgColor.rgb
        and fill.fgColor.rgb.upper() in HEADLINE_COLORS
    )


def _apply_cell_style(cell, style_cfg: dict) -> None:
    font_color = style_cfg.get("font_color")
    font_bold = style_cfg.get("font_bold")
    if font_color or font_bold is not None:
        new_font = copy(cell.font)
        if font_color:
            rgb = str(font_color).replace("#", "").upper()
            if len(rgb) == 6:
                rgb = f"FF{rgb}"
            new_font.color = Color(rgb=rgb)
        if font_bold is not None:
            new_font.bold = bool(font_bold)
        cell.font = new_font

    number_format = style_cfg.get("number_format")
    if number_format:
        cell.number_format = str(number_format)

    align_horizontal = style_cfg.get("align_horizontal")
    align_vertical = style_cfg.get("align_vertical")
    wrap_text = style_cfg.get("wrap_text")
    if align_horizontal or align_vertical or wrap_text is not None:
        new_alignment = copy(cell.alignment)
        if align_horizontal:
            new_alignment.horizontal = str(align_horizontal)
        if align_vertical:
            new_alignment.vertical = str(align_vertical)
        if wrap_text is not None:
            new_alignment.wrap_text = bool(wrap_text)
        cell.alignment = new_alignment


def _fill_static_sheet(ws, cell_map: dict, transformed_data: dict) -> int:
    filled = 0
    for field_key, cell_ref in cell_map.items():
        cell = ws[cell_ref]
        if _is_headline_cell(cell):
            continue
        cell.value = transformed_data.get(field_key, "")
        filled += 1
    return filled


def _fill_dynamic_sheet(ws, transformed_data: dict) -> int:
    filled = 0

    cell_styles = {}
    if isinstance(transformed_data, dict):
        raw_styles = transformed_data.pop("_cell_styles", {})
        if isinstance(raw_styles, dict):
            cell_styles = raw_styles

    for cell_ref, value in transformed_data.items():
        if not (isinstance(cell_ref, str) and len(cell_ref) >= 2):
            continue
        try:
            cell = ws[cell_ref]
            if _is_headline_cell(cell):
                continue
            cell.value = value
            _apply_cell_style(cell, cell_styles.get(cell_ref, {}))
            filled += 1
        except Exception as e:
            st.error(f"Error filling {cell_ref}: {e}")
    return filled


def fill_excel(template_bytes, field_values, summary_text, return_report=False):
    """
    Fill Excel template with data from field_values.

    Args:
        template_bytes: Excel template file as bytes
        field_values: Dictionary of field values to fill
        summary_text: Company summary text
        return_report: If True, return (excel_bytes, report)

    Returns:
        Filled Excel file as bytes
        OR tuple(excel_bytes, report) if return_report=True
    """
    # Keep partial rich-text formatting in template header cells
    # (e.g., "Forsikringssum (kr) / Premium" where only "Premium" is blue).
    try:
        wb = load_workbook(filename=BytesIO(template_bytes), rich_text=True)
    except TypeError:
        wb = load_workbook(filename=BytesIO(template_bytes))

    # Debug visibility; main_page suppresses these in production mode.
    st.write("EXCEL_FILLER DEBUG")
    st.write(f"field_values keys: {list(field_values.keys())}")
    if "pdf_text" in field_values:
        st.success(f"pdf_text exists; length={len(field_values['pdf_text'])}")
    else:
        st.error("pdf_text not found in field_values")

    report = {
        "sheets": [],
        "summary": {"status": "not_run", "cell": None, "error": None},
    }

    for sheet_name in SHEET_MAPPINGS.keys():
        if sheet_name not in wb.sheetnames:
            st.warning(f"Sheet '{sheet_name}' not found in template")
            report["sheets"].append(
                {
                    "sheet": sheet_name,
                    "status": "missing_in_template",
                    "mapping_type": None,
                    "filled_cells": 0,
                    "error": None,
                }
            )
            continue

        ws = wb[sheet_name]
        sheet_result = {
            "sheet": sheet_name,
            "status": "ok",
            "mapping_type": None,
            "filled_cells": 0,
            "error": None,
        }

        try:
            transformed_data = transform_for_sheet(sheet_name, field_values)
            cell_map = SHEET_MAPPINGS.get(sheet_name, {})

            if cell_map:
                sheet_result["mapping_type"] = "static"
                sheet_result["filled_cells"] = _fill_static_sheet(ws, cell_map, transformed_data)
            else:
                sheet_result["mapping_type"] = "dynamic"
                sheet_result["filled_cells"] = _fill_dynamic_sheet(ws, transformed_data)
        except Exception as e:
            sheet_result["status"] = "failed"
            sheet_result["error"] = str(e)
            st.error(f"Sheet failed (continuing): {sheet_name} -> {e}")

        report["sheets"].append(sheet_result)

    # Summary placement is also fail-soft.
    first_sheet = wb.sheetnames[0]
    ws_first = wb[first_sheet]
    if summary_text:
        placed = False
        try:
            for row in ws_first.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and "skriv her" in cell.value.lower():
                        cell.value = summary_text
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                        report["summary"] = {"status": "ok", "cell": cell.coordinate, "error": None}
                        placed = True
                        break
                if placed:
                    break

            if not placed:
                ws_first["A46"] = summary_text
                ws_first["A46"].alignment = Alignment(wrap_text=True, vertical="top")
                report["summary"] = {"status": "fallback", "cell": "A46", "error": None}
        except Exception as e:
            report["summary"] = {"status": "failed", "cell": None, "error": str(e)}
            st.error(f"Summary placement failed: {e}")

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    excel_bytes = out.getvalue()
    st.success("Excel file filled successfully")

    if return_report:
        return excel_bytes, report
    return excel_bytes


def run():
    """Streamlit page view for the excel filler module"""
    st.title("Excel Filler Module")
    st.write("This module fills Excel templates with extracted data.")
    st.info("Used by the main page to create filled Excel files.")

